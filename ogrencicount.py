import openpyxl
import pdb


wb_obj = openpyxl.load_workbook("untitled2ogrenci.xlsx")
#print(wb_obj)
sheet_obj = wb_obj.active
#print(sheet_obj)
#cell_obj = sheet_obj.cell(row=1, column=1)
#print(cell_obj)
#print(cell_obj.value)
max_row = sheet_obj.max_row
#sheet_obj["A1"] = 2
#wb_obj.save("ogrenci.xlsx")


'''
i = 1

while(i <= max_row):
    i = i + 1
    fakulte = sheet_obj.cell(row=i, column=7).value
    if not fakulte:
        print(i)
        continue
    if fakulte.strip() == 'TIP':
        print(sheet_obj.cell(row=i, column=1).value)
'''



'''
i = 2
indeks = 2
while(i <= 100):
    adsoyad = sheet_obj.cell(row=i, column=4).value
    sheet_obj.cell(row=indeks, column=5).value = adsoyad
    tarih = sheet_obj.cell(row=i, column=3).value
    sheet_obj.cell(row=indeks, column=6).value = tarih
    indeks = indeks + 1
    a = 7
    while adsoyad == sheet_obj.cell(row=i, column=4).value:
        konu = sheet_obj.cell(row=i, column=1).value
        sheet_obj.cell(row=indeks-1, column=a).value = konu
        a = a + 1
        i = i + 1
'''
'''
for i in range(1, max_row+1):
    if sheet_obj.cell(row=i, column=2).value[0].isnumeric():
        sheet_obj.cell(row=i, column=1).value = ""


wb_obj.save("countogrenci.xlsx")
'''


#def bubbleSort(arr):




turler = ["Uygarlık Tarihi", "Din", "Selçuklular", "ABD Tarihi", "Latin Amerika Tarihi", "İktisat Bilimi", "Politik Bilimler",
          "Ticaret Hukuku", "Eğitim Bilimleri", "Müzik", "Sanat Tarihi", "Medya", "Mitoloji", "Fransız Edebiyatı",
          "İngiliz Edebiyatı", "Alman Edebiyatı", "Rus Edebiyatı", "Macar Edebiyatı", "Amerikan Edebiyatı", "Mısır Edebiyatı",
          "İran Edebiyatı", "Türk Edebiyatı", "Eleştiri", "Bilim Tarihi", "Matematik", "Astronomi", "Fizik", "Kimya", "Jeoloji",
          "Tıp", "Cumhuriyet Dönemi Ansiklopedileri", "Veterinerlik", "Mühendislik Ders Kitapları", "Askerlik Sanatı",
          "Mühendislik", "Kütüphaneler"]

fakulteler = ["CUMHURİYET MYO", "DİŞ HEKİMLİĞİ", "ECZACILIK", "EDEBİYAT", "EĞİTİM", "EĞİTİM BİLİMLERİ ENSTİTÜSÜ",
              "FEN BİLİMLERİ ENSTİTÜSÜ", "FEN", "İLAHİYAT", "İLETİŞİM", "MİMARLIK", "GÜZEL SANATLAR", "MÜHENDİSLİK",
              "SAĞLIK BİLİMLERİ", "SAĞLIK BİLİMLERİ ENSTİTÜSÜ", "SAĞLIK HİZMETLERİ MYO", "SİVAS MYO", "SOSYAL BİLİMLER ENSTİTÜSÜ",
              "SUŞEHRİ SAĞLIK YÜKSEKOKULU", "ŞARKIŞLA AŞIK VEYSEL MYO", "TEKNOLOJİ", "TIP", "TURİZM", "VETERİNER"]

sirala = []
a = 0
while(a <= 23):
    x = 0
    print(fakulteler[a], "-------------------------------------")
    while(x <= 35):
        i = 1
        sayac = 0
        while(i <= max_row):
            i = i + 1
            fakulte = sheet_obj.cell(row=i, column=5).value
            if not fakulte:
                continue
            if fakulte.strip() == fakulteler[a] and sheet_obj.cell(row=i, column=1).value == turler[x]:
                sayac = sayac + 1
            if i == max_row:
                sirala.append(turler[x])
                sirala.append(sayac)
                #sirala[sira] = turler[x]
                #sira = sira + 1
                #sirala[sira] = sayac
                #print(turler[x], " >> ", sayac)
                x = x + 1
    #for x in range(len(sirala)):
        #print(sirala[x])
    #pdb.set_trace()
    n = len(sirala)
    # Traverse through all array elements
    for i in range(n+1):

        # Last i elements are already in place
        for j in range(0, n - i - 3, 2):
            # traverse the array from 0 to n-i-1
            # Swap if the element found is greater
            # than the next element
            #print("sirala[j] = ", " >> ", sirala[j])
            #print("sirala[j + 1] = ", " >> ", sirala[j + 1])
            #print("sirala[j + 2] = ", " >> ", sirala[j + 2])
            #print("sirala[j + 3] = ", " >> ", sirala[j + 3])
            if sirala[j+1] < sirala[j + 3]:
                temp0 = sirala[j]
                temp1 = sirala[j+1]
                temp2 = sirala[j+2]
                temp3 = sirala[j+3]
                sirala[j] = temp2
                sirala[j+1] = temp3
                sirala[j+2] = temp0
                sirala[j+3] = temp1
                #sirala[j], sirala[j+1], sirala[j+2], sirala[j+3] = sirala[j+2], sirala[j+3], sirala[j], sirala[j+1]

    for x in range(0,len(sirala),2):
        print(sirala[x], " > ", sirala[x+1])
    a = a + 1

    sirala.clear()


#wb_obj.save("untitled2ogrenci.xlsx")


#https://www.geeksforgeeks.org/python-program-for-bubble-sort/