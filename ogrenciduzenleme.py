import openpyxl
import pdb

wb_obj = openpyxl.load_workbook("yeniogrenci.xlsx")
#print(wb_obj)
sheet_obj = wb_obj.active
#print(sheet_obj)
#cell_obj = sheet_obj.cell(row=1, column=1)
#print(cell_obj)
#print(cell_obj.value)
max_row = sheet_obj.max_row
#sheet_obj["A1"] = 2
#wb_obj.save("ogrenci.xlsx")
#pdb.set_trace()

'''
for i in range(1, max_row+1):
    if sheet_obj.cell(row = i, column=2).value[0] == "A" or sheet_obj.cell(row = i, column=2).value[0] == "C": sheet_obj.cell(row=i, column=1).value = "Uygarlık Tarihi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "B": sheet_obj.cell(row=i, column=1).value = "Din"
    elif sheet_obj.cell(row = i, column=2).value[0] == "D": sheet_obj.cell(row=i, column=1).value = "Selçuklular"
    elif sheet_obj.cell(row = i, column=2).value[0] == "E": sheet_obj.cell(row=i, column=1).value = "ABD Tarihi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "F": sheet_obj.cell(row=i, column=1).value = "Latin Amerika Tarihi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "H": sheet_obj.cell(row=i, column=1).value = "İktisat Bilimi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "J": sheet_obj.cell(row=i, column=1).value = "Politik Bilimler"
    elif sheet_obj.cell(row = i, column=2).value[0] == "K": sheet_obj.cell(row=i, column=1).value = "Ticaret Hukuku"
    elif sheet_obj.cell(row = i, column=2).value[0] == "L": sheet_obj.cell(row=i, column=1).value = "Eğitim Bilimleri"
    elif sheet_obj.cell(row = i, column=2).value[0] == "M": sheet_obj.cell(row=i, column=1).value = "Müzik"
    elif sheet_obj.cell(row = i, column=2).value[0] == "N": sheet_obj.cell(row=i, column=1).value = "Sanat Tarihi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "P" and sheet_obj.cell(row=i, column=2).value[1] == " ": sheet_obj.cell(row=i, column=1).value = "Medya"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PA": sheet_obj.cell(row=i, column=1).value = "Mitoloji"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PC" or sheet_obj.cell(row=i, column=2).value[0:2] == "PQ": sheet_obj.cell(row=i, column=1).value = "Fransız Edebiyatı"
    elif sheet_obj.cell(row = i, column=2).value[0:2] == "PE" or sheet_obj.cell(row = i, column=2).value[0:2] == "PR": sheet_obj.cell(row=i, column=1).value = "İngiliz Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PF" or sheet_obj.cell(row = i, column=2).value[0:2] == "PT": sheet_obj.cell(row=i, column=1).value = "Alman Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PG": sheet_obj.cell(row=i, column=1).value = "Rus Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PH": sheet_obj.cell(row=i, column=1).value = "Macar Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PS": sheet_obj.cell(row=i, column=1).value = "Amerikan Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PJ": sheet_obj.cell(row=i, column=1).value = "Mısır Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PK": sheet_obj.cell(row=i, column=1).value = "İran Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PL": sheet_obj.cell(row=i, column=1).value = "Türk Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PN": sheet_obj.cell(row=i, column=1).value = "Eleştiri"
    elif sheet_obj.cell(row=i, column=2).value[0] == "Q" and sheet_obj.cell(row=i, column=2).value[1] == " ": sheet_obj.cell(row=i, column=1).value = "Bilim Tarihi"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QA": sheet_obj.cell(row=i, column=1).value = "Matematik"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QB": sheet_obj.cell(row=i, column=1).value = "Astronomi"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QC": sheet_obj.cell(row=i, column=1).value = "Fizik"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QD": sheet_obj.cell(row=i, column=1).value = "Kimya"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QE": sheet_obj.cell(row=i, column=1).value = "Jeoloji"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QH" or sheet_obj.cell(row = i, column=2).value[0:2] == "QK" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QL" or sheet_obj.cell(row = i, column=2).value[0:2] == "QM" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QP" or sheet_obj.cell(row = i, column=2).value[0:2] == "QR" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QS" or sheet_obj.cell(row = i, column=2).value[0:2] == "QT" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QV" or sheet_obj.cell(row = i, column=2).value[0:2] == "QU" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QW" or sheet_obj.cell(row = i, column=2).value[0:2] == "QY" or \
    sheet_obj.cell(row=i, column=2).value[0:2] == "QX" or sheet_obj.cell(row = i, column=2).value[0:2] == "QZ" or \
    sheet_obj.cell(row=i, column=2).value[0] == "W" : sheet_obj.cell(row=i, column=1).value = "Tıp"
    elif sheet_obj.cell(row = i, column=2).value[0] == "R": sheet_obj.cell(row=i, column=1).value = "Cumhuriyet Dönemi Ansiklopedileri"
    elif sheet_obj.cell(row = i, column=2).value[0] == "S": sheet_obj.cell(row=i, column=1).value = "Veterinerlik"
    elif sheet_obj.cell(row=i, column=2).value[0] == "T": sheet_obj.cell(row=i, column=1).value = "Mühendislik Ders Kitapları"
    elif sheet_obj.cell(row=i, column=2).value[0] == "U": sheet_obj.cell(row=i, column=1).value = "Askerlik Sanatı"
    elif sheet_obj.cell(row = i, column=2).value[0] == "V": sheet_obj.cell(row=i, column=1).value = "Mühendislik"
    elif sheet_obj.cell(row=i, column=2).value[0] == "Z": sheet_obj.cell(row=i, column=1).value = "Kütüphaneler"

wb_obj.save("ogrenci.xlsx")
'''

i = 2
indeks = 2
while(i <= max_row):
    adsoyad = sheet_obj.cell(row=i, column=4).value
    sheet_obj.cell(row=indeks, column=5).value = adsoyad
    tarih = sheet_obj.cell(row=i, column=3).value
    sheet_obj.cell(row=indeks, column=6).value = tarih
    indeks = indeks + 1
    a = 7
    while adsoyad == sheet_obj.cell(row=i, column=4).value:
        konu = sheet_obj.cell(row=i, column=1).value
        sheet_obj.cell(row=indeks-1, column=a).value = konu
        a = a + 1
        i = i + 1

'''
i = 1
while(i <= max_row):
    i = i + 1
    a = 7
    while(sheet_obj.cell(row=i, column=a).value != None):
        kayit = sheet_obj.cell(row=i, column=a).value
        x = a + 1
        while(sheet_obj.cell(row=i, column=x).value != None):
            #print(sheet_obj.cell(row=i, column=x).value)
            if kayit == sheet_obj.cell(row=i, column=x).value:
                sheet_obj.cell(row=i, column=x).value = ""
                #x = x + 1
            x = x + 1
        a = a + 1
'''
wb_obj.save("yeniogrenci.xlsx")




i = 2
indeks = 2
while(i <= max_row):
    adsoyad = sheet_obj.cell(row=i, column=4).value
    sheet_obj.cell(row=indeks, column=6).value = adsoyad
    tarih = sheet_obj.cell(row=i, column=3).value
    sheet_obj.cell(row=indeks, column=7).value = tarih
    indeks = indeks + 1
    a = 8
    while adsoyad == sheet_obj.cell(row=i, column=4).value:
        konu = sheet_obj.cell(row=i, column=1).value
        sheet_obj.cell(row=indeks-1, column=a).value = konu
        a = a + 1
        i = i + 1



wb_obj.save("yenideneme.xlsx")


'''
# arama fonksiyonu
a = 1
for i in range(1, max_row+1):
    if sheet_obj.cell(row=i, column=2).value[0:2] == "GF":
        print("var ", a)
        a += 1
    else: print("yok")
    a += 1
'''

#if sheet_obj.cell(row=i, column=2).value[0] == "Q" and sheet_obj.cell(row=i, column=2).value[1] == " ":
#if sheet_obj.cell(row=i, column=2).value[0].isnumeric():


'''
# boşluk haline getirme fonksiyonu

for i in range(1, max_row+1):
    if sheet_obj.cell(row=i, column=2).value[0:2] == "PZ":
        sheet_obj.cell(row=i, column=1).value = ""


wb_obj.save("ogrenci.xlsx")
'''

'''

turler = ["Uygarlık Tarihi", "Din", "Selçuklular", "ABD Tarihi", "Latin Amerika Tarihi", "İktisat Bilimi", "Politik Bilimler",
          "Ticaret Hukuku", "Eğitim Bilimleri", "Müzik", "Sanat Tarihi", "Medya", "Mitoloji", "Fransız Edebiyatı",
          "İngiliz Edebiyatı", "Alman Edebiyatı", "Rus Edebiyatı", "Macar Edebiyatı", "Amerikan Edebiyatı", "Mısır Edebiyatı",
          "İran Edebiyatı", "Türk Edebiyatı", "Eleştiri", "Bilim Tarihi", "Matematik", "Astronomi", "Fizik", "Kimya", "Jeoloji",
          "Tıp", "Cumhuriyet Dönemi Ansiklopedileri", "Veterinerlik", "Mühendislik Ders Kitapları", "Askerlik Sanatı",
          "Mühendislik", "Kütüphaneler"]

fakulteler = ["CUMHURİYET MYO", "DİŞ HEKİMLİĞİ", "ECZACILIK", "EDEBİYAT", "EĞİTİM", "EĞİTİM BİLİMLERİ ENSTİTÜSÜ",
              "FEN BİLİMLERİ ENSTİTÜSÜ", "FEN", "İLAHİYAT", "İLETİŞİM", "MİMARLIK", "GÜZEL SANATLAR", "MÜHENDİSLİK",
              "SAĞLIK BİLİMLERİ", "SAĞLIK BİLİMLERİ ENSTİTÜSÜ", "SAĞLIK HİZMETLERİ MYO", "SİVAS MYO", "SOSYAL BİLİMLER ENSTİTÜSÜ",
              "SUŞEHRİ SAĞLIK YÜKSEKOKULU", "ŞARKIŞLA AŞIK VEYSEL MYO", "TEKNOLOJİ", "TIP", "TURİZM", "VETERİNER"]

a = 0
while(a <= 23):
    x = 0
    print(fakulteler[a], "-------------------------------------")
    while(x <= 35):
        i = 1
        sayac = 0
        while(i <= max_row):
            i = i + 1
            fakulte = sheet_obj.cell(row=i, column=5).value
            if not fakulte:
                continue
            if fakulte.strip() == fakulteler[a] and sheet_obj.cell(row=i, column=1).value == turler[x]:
                sayac = sayac + 1
            if i == max_row:
                print(turler[x], " >> ", sayac)
                x = x + 1
                '''
    #a = a + 1

'''
#çokluları silme kodu
#pdb.set_trace()
i = 1
while(i <= max_row):
    i = i + 1
    a = 8
    while(sheet_obj.cell(row=i, column=a).value != None):
        kayit = sheet_obj.cell(row=i, column=a).value
        x = a + 1
        while(sheet_obj.cell(row=i, column=x).value != None):
            #print(sheet_obj.cell(row=i, column=x).value)
            if kayit == sheet_obj.cell(row=i, column=x).value:
                sheet_obj.cell(row=i, column=x).value = ""
                #x = x + 1
            x = x + 1
        a = a + 1

wb_obj.save("ogrenci.xlsx")
'''

#teklileri boşluk haline getirme kodu
i = 1
while(i <= max_row):
    i = i + 1
    if sheet_obj.cell(row=i, column=9).value == None:
        sheet_obj.cell(row=i, column=8).value = ""

wb_obj.save("ogrenci.xlsx")
