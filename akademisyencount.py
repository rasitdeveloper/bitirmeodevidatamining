import openpyxl
import pdb

wb_obj = openpyxl.load_workbook("akademisyencount.xlsx")
#print(wb_obj)
sheet_obj = wb_obj.active
#print(sheet_obj)
#cell_obj = sheet_obj.cell(row=1, column=1)
#print(cell_obj)
#print(cell_obj.value)
max_row = sheet_obj.max_row
#sheet_obj["A1"] = 2
#wb_obj.save("ogrenci.xlsx")
#pdb.set_trace()

'''
for i in range(1, max_row+1):
    if sheet_obj.cell(row = i, column=2).value[0] == "A" or sheet_obj.cell(row = i, column=2).value[0] == "C": sheet_obj.cell(row=i, column=1).value = "Uygarlık Tarihi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "B": sheet_obj.cell(row=i, column=1).value = "Din"
    elif sheet_obj.cell(row = i, column=2).value[0] == "D": sheet_obj.cell(row=i, column=1).value = "Selçuklular"
    elif sheet_obj.cell(row = i, column=2).value[0] == "E": sheet_obj.cell(row=i, column=1).value = "ABD Tarihi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "F": sheet_obj.cell(row=i, column=1).value = "Latin Amerika Tarihi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "H": sheet_obj.cell(row=i, column=1).value = "İktisat Bilimi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "J": sheet_obj.cell(row=i, column=1).value = "Politik Bilimler"
    elif sheet_obj.cell(row = i, column=2).value[0] == "K": sheet_obj.cell(row=i, column=1).value = "Ticaret Hukuku"
    elif sheet_obj.cell(row = i, column=2).value[0] == "L": sheet_obj.cell(row=i, column=1).value = "Eğitim Bilimleri"
    elif sheet_obj.cell(row = i, column=2).value[0] == "M": sheet_obj.cell(row=i, column=1).value = "Müzik"
    elif sheet_obj.cell(row = i, column=2).value[0] == "N": sheet_obj.cell(row=i, column=1).value = "Sanat Tarihi"
    elif sheet_obj.cell(row = i, column=2).value[0] == "P" and sheet_obj.cell(row=i, column=2).value[1] == " ": sheet_obj.cell(row=i, column=1).value = "Medya"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PA": sheet_obj.cell(row=i, column=1).value = "Mitoloji"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PC" or sheet_obj.cell(row=i, column=2).value[0:2] == "PQ": sheet_obj.cell(row=i, column=1).value = "Fransız Edebiyatı"
    elif sheet_obj.cell(row = i, column=2).value[0:2] == "PE" or sheet_obj.cell(row = i, column=2).value[0:2] == "PR": sheet_obj.cell(row=i, column=1).value = "İngiliz Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PF" or sheet_obj.cell(row = i, column=2).value[0:2] == "PT": sheet_obj.cell(row=i, column=1).value = "Alman Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PG": sheet_obj.cell(row=i, column=1).value = "Rus Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PH": sheet_obj.cell(row=i, column=1).value = "Macar Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PS": sheet_obj.cell(row=i, column=1).value = "Amerikan Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PJ": sheet_obj.cell(row=i, column=1).value = "Mısır Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PK": sheet_obj.cell(row=i, column=1).value = "İran Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PL": sheet_obj.cell(row=i, column=1).value = "Türk Edebiyatı"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "PN": sheet_obj.cell(row=i, column=1).value = "Eleştiri"
    elif sheet_obj.cell(row=i, column=2).value[0] == "Q" and sheet_obj.cell(row=i, column=2).value[1] == " ": sheet_obj.cell(row=i, column=1).value = "Bilim Tarihi"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QA": sheet_obj.cell(row=i, column=1).value = "Matematik"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QB": sheet_obj.cell(row=i, column=1).value = "Astronomi"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QC": sheet_obj.cell(row=i, column=1).value = "Fizik"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QD": sheet_obj.cell(row=i, column=1).value = "Kimya"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QE": sheet_obj.cell(row=i, column=1).value = "Jeoloji"
    elif sheet_obj.cell(row=i, column=2).value[0:2] == "QH" or sheet_obj.cell(row = i, column=2).value[0:2] == "QK" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QL" or sheet_obj.cell(row = i, column=2).value[0:2] == "QM" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QP" or sheet_obj.cell(row = i, column=2).value[0:2] == "QR" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QS" or sheet_obj.cell(row = i, column=2).value[0:2] == "QT" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QV" or sheet_obj.cell(row = i, column=2).value[0:2] == "QU" or \
    sheet_obj.cell(row = i, column=2).value[0:2] == "QW" or sheet_obj.cell(row = i, column=2).value[0:2] == "QY" or \
    sheet_obj.cell(row=i, column=2).value[0:2] == "QX" or sheet_obj.cell(row = i, column=2).value[0:2] == "QZ" or \
    sheet_obj.cell(row=i, column=2).value[0] == "W" : sheet_obj.cell(row=i, column=1).value = "Tıp"
    elif sheet_obj.cell(row = i, column=2).value[0] == "R": sheet_obj.cell(row=i, column=1).value = "Cumhuriyet Dönemi Ansiklopedileri"
    elif sheet_obj.cell(row = i, column=2).value[0] == "S": sheet_obj.cell(row=i, column=1).value = "Veterinerlik"
    elif sheet_obj.cell(row=i, column=2).value[0] == "T": sheet_obj.cell(row=i, column=1).value = "Mühendislik Ders Kitapları"
    elif sheet_obj.cell(row=i, column=2).value[0] == "U": sheet_obj.cell(row=i, column=1).value = "Askerlik Sanatı"
    elif sheet_obj.cell(row = i, column=2).value[0] == "V": sheet_obj.cell(row=i, column=1).value = "Mühendislik"
    elif sheet_obj.cell(row=i, column=2).value[0] == "Z": sheet_obj.cell(row=i, column=1).value = "Kütüphaneler"

wb_obj.save("akademisyencount.xlsx")
'''

turler = ["Uygarlık Tarihi", "Din", "Selçuklular", "İktisat Bilimi", "Politik Bilimler", "Ticaret Hukuku",
          "Eğitim Bilimleri", "Müzik", "Sanat Tarihi", "Medya", "Mitoloji", "Fransız Edebiyatı",
          "İngiliz Edebiyatı", "Alman Edebiyatı", "Rus Edebiyatı", "Amerikan Edebiyatı", "Mısır Edebiyatı",
          "Türk Edebiyatı", "Eleştiri", "Bilim Tarihi", "Matematik", "Fizik", "Jeoloji",
          "Tıp", "Cumhuriyet Dönemi Ansiklopedileri", "Veterinerlik", "Mühendislik Ders Kitapları", "Askerlik Sanatı",
          "Mühendislik", "Kütüphaneler"]

fakulteler = ["YILDIZELİ MYO", "CUMHURİYET MYO", "SUŞEHRİ MYO", "GENEL SEKRETERLİK", "FEN FAKÜLTESİ", "SİVAS MYO",
              "EĞİTİM FAKÜLTESİ", "REKTÖRLÜK", "EDEBİYAT FAKÜLTESİ", "FEN BİLİMLERİ ENSTİTÜSÜ", "İLETİŞİM FAKÜLTESİ",
              "HAFİK KAMER ÖRNEK MYO", "SAĞLIK BİLİMLERİ FAKÜLTESİ", "TIP FAKÜLTESİ", "ZARA VEYSEL DURSUN UY.BİL.Y.O.",
              "MÜHENDİSLİK FAKÜLTESİ", "SUŞEHRİ TİMUR KARABAL MYO", "GÜRÜN MYO", "TURİZM İŞL. VE OTELCİLİK Y.O.",
              "SAĞLIK HİZMETLERİ MYO", "FEN-EDEBİYAT FAKÜLTESİ", "KANGAL MYO", "SOSYAL BİLİMLER ENSTİTÜSÜ",
              "ENFORMATİK BÖLÜMÜ", "ZARA MYO", "TURİZM FAKÜLTESİ", "İKTİSADİ VE İDARİ BİLİMLER FAKÜLTESİ",
              "İLETİŞİM FAKÜLTESİ", "TEKNOLOJİ FAKÜLTESİ", "BEDEN EĞİTİMİ VE SPOR YO.", "SAĞLIK BİLİMLERİ ENSTİTÜSÜ",
              "İLAHİYAT FAKÜLTESİ", "GÜZEL SANATLAR FAKÜLTESİ", "ZARA AHMET ÇUHADAROĞLU MYO", "YABANCI DILLER YÜKSEKOKULU",
              "ECZACILIK FAKÜLTESİ", "VETERİNER FAKÜLTESİ", "MİMARLIK FAKÜLTESİ"]
'''
a = 0
while (a <= 36):
    x = 0
    print(fakulteler[a], "-------------------------------------")
    while (x <= 29):
        i = 1
        sayac = 0
        while (i <= max_row):
            i = i + 1
            fakulte = sheet_obj.cell(row=i, column=5).value
            if not fakulte:
                continue
            if fakulte.strip() == fakulteler[a] and sheet_obj.cell(row=i, column=1).value == turler[x]:
                sayac = sayac + 1
            if i == max_row:
                print(turler[x], " >> ", sayac)
                x = x + 1
    a = a + 1

#wb_obj.save("akademisyencount.xlsx")
'''

sirala = []
a = 0
while(a <= 37):
    x = 0
    print(fakulteler[a], "-------------------------------------")
    while(x <= 29):
        i = 1
        sayac = 0
        while(i <= max_row):
            i = i + 1
            fakulte = sheet_obj.cell(row=i, column=5).value
            if not fakulte:
                continue
            if fakulte.strip() == fakulteler[a] and sheet_obj.cell(row=i, column=1).value == turler[x]:
                sayac = sayac + 1
            if i == max_row:
                sirala.append(turler[x])
                sirala.append(sayac)
                #sirala[sira] = turler[x]
                #sira = sira + 1
                #sirala[sira] = sayac
                #print(turler[x], " >> ", sayac)
                x = x + 1
    #for x in range(len(sirala)):
        #print(sirala[x])
    #pdb.set_trace()
    n = len(sirala)
    # Traverse through all array elements
    for i in range(n+1):

        # Last i elements are already in place
        for j in range(0, n - i - 3, 2):
            # traverse the array from 0 to n-i-1
            # Swap if the element found is greater
            # than the next element
            #print("sirala[j] = ", " >> ", sirala[j])
            #print("sirala[j + 1] = ", " >> ", sirala[j + 1])
            #print("sirala[j + 2] = ", " >> ", sirala[j + 2])
            #print("sirala[j + 3] = ", " >> ", sirala[j + 3])
            if sirala[j+1] < sirala[j + 3]:
                temp0 = sirala[j]
                temp1 = sirala[j+1]
                temp2 = sirala[j+2]
                temp3 = sirala[j+3]
                sirala[j] = temp2
                sirala[j+1] = temp3
                sirala[j+2] = temp0
                sirala[j+3] = temp1
                #sirala[j], sirala[j+1], sirala[j+2], sirala[j+3] = sirala[j+2], sirala[j+3], sirala[j], sirala[j+1]

    for x in range(0, len(sirala), 2):
        print(sirala[x], " > ", sirala[x + 1])
    a = a + 1

    sirala.clear()